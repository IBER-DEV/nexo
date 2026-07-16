from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from django.db import transaction
from django.db.models import Q
from openpyxl import load_workbook
from .models import Activity, Empresa, Proceso, Aplicacion
from .serializers import ActivitySerializer
from .filters import ActivityFilter
from .sync_utils import normalize_header, parse_date, parse_pk, get_or_create_responsable


class ActivityViewSet(viewsets.ModelViewSet):
    queryset = Activity.objects.select_related("responsable", "created_by").order_by("-pk")
    serializer_class = ActivitySerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_class = ActivityFilter
    search_fields = ["nombre", "empresa", "aplicacion", "responsable__nombre", "mes_planeacion"]
    ordering_fields = [
        "pk",
        "nombre",
        "prioridad",
        "estado",
        "fecha_limite",
        "fecha_inicio",
        "mes_planeacion",
        "semana_planeacion",
    ]
    http_method_names = ["get", "post", "patch", "delete", "head", "options"]

    def get_queryset(self):
        base_qs = Activity.objects.select_related("responsable", "created_by").order_by("-pk")
        user = self.request.user
        if getattr(user, "is_admin", False):
            return base_qs
        if getattr(user, "is_coordinator", False):
            team_ids = user.team_user_ids() if hasattr(user, "team_user_ids") else [user.pk]
            return base_qs.filter(Q(responsable_id__in=team_ids) | Q(created_by_id__in=team_ids)).distinct()
        return base_qs.filter(Q(responsable=user) | Q(created_by=user)).distinct()

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=["get"], url_path="meta")
    def meta(self, request):
        is_admin = getattr(request.user, "is_admin", False)
        activities = self.get_queryset()
        if is_admin:
            empresas = Empresa.objects.values_list("nombre", flat=True).order_by("nombre")
            procesos = Proceso.objects.values_list("nombre", flat=True).order_by("nombre")
            aplicaciones = Aplicacion.objects.values_list("nombre", flat=True).order_by("nombre")
        else:
            empresas = activities.values_list("empresa", flat=True).distinct().order_by("empresa")
            procesos = activities.values_list("proceso", flat=True).distinct().order_by("proceso")
            aplicaciones = activities.values_list("aplicacion", flat=True).distinct().order_by("aplicacion")
        payload = {
            "empresas": list(empresas),
            "procesos": list(procesos),
            "aplicaciones": list(aplicaciones),
            "stakeholders": list(
                activities.exclude(stakeholder="")
                .values_list("stakeholder", flat=True)
                .distinct()
                .order_by("stakeholder")
            ),
        }
        return Response(payload)

    @action(
        detail=False,
        methods=["post"],
        url_path="import",
        parser_classes=[MultiPartParser, FormParser],
    )
    def import_excel(self, request):
        upload = request.FILES.get("file")
        if not upload:
            return Response({"detail": "Archivo requerido"}, status=status.HTTP_400_BAD_REQUEST)

        mes_planeacion = (request.data.get("mes_planeacion") or "").strip() or None
        semana_planeacion_raw = (request.data.get("semana_planeacion") or "").strip()
        semana_planeacion = int(semana_planeacion_raw) if semana_planeacion_raw.isdigit() else None
        dry_run = str(request.data.get("dry_run") or "").lower() in {"1", "true", "yes"}

        if semana_planeacion is not None and not 1 <= semana_planeacion <= 5:
            return Response({"detail": "Semana invalida"}, status=status.HTTP_400_BAD_REQUEST)

        expected_headers = {
            "empresa": "empresa",
            "proceso": "proceso",
            "aplicacion": "aplicacion",
            "nombre actividad": "nombre",
            "descripcion actividad": "descripcion",
            "responsable": "responsable",
            "stakeholder": "stakeholder",
            "fecha inicio": "fecha_inicio",
            "fecha finalizacion": "fecha_limite",
        }
        optional_id_headers = {"id", "codigo", "codigo actividad", "id actividad"}

        wb = load_workbook(upload, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({"detail": "Archivo sin filas"}, status=status.HTTP_400_BAD_REQUEST)

        header_row_index = None
        header_row = None
        for idx, row in enumerate(rows, start=1):
            values = [normalize_header(v) for v in row]
            matches = set(expected_headers.keys()).intersection(values)
            if matches:
                header_row_index = idx
                header_row = row
                if len(matches) >= 4:
                    break

        if header_row_index is None or header_row is None:
            return Response({"detail": "No se encontro fila de encabezados"}, status=status.HTTP_400_BAD_REQUEST)

        header_map: dict[str, int] = {}
        id_index: int | None = None
        for idx, header in enumerate(header_row):
            key = normalize_header(header)
            if key in expected_headers:
                header_map[expected_headers[key]] = idx
            if key in optional_id_headers:
                id_index = idx

        missing = [k for k in expected_headers.values() if k not in header_map]
        if missing:
            return Response(
                {"detail": "Faltan columnas requeridas", "missing": missing},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created = 0
        updated = 0
        skipped = 0
        errors: list[dict[str, object]] = []

        with transaction.atomic():
            for row_index, row in enumerate(rows[header_row_index:], start=header_row_index + 1):
                if not any(cell not in (None, "") for cell in row):
                    continue

                empresa = row[header_map["empresa"]]
                proceso = row[header_map["proceso"]]
                aplicacion = row[header_map["aplicacion"]]
                nombre = row[header_map["nombre"]]
                descripcion = row[header_map["descripcion"]]
                responsable_name = row[header_map["responsable"]]
                stakeholder = row[header_map["stakeholder"]]
                fecha_inicio = parse_date(row[header_map["fecha_inicio"]])
                fecha_limite = parse_date(row[header_map["fecha_limite"]])

                if not (empresa and proceso and aplicacion and nombre and responsable_name and fecha_inicio and fecha_limite):
                    skipped += 1
                    errors.append({"row": row_index, "error": "Campos requeridos incompletos"})
                    continue

                responsable = get_or_create_responsable(responsable_name, requesting_user=request.user)

                data = {
                    "empresa": str(empresa).strip(),
                    "proceso": str(proceso).strip(),
                    "aplicacion": str(aplicacion).strip(),
                    "nombre": str(nombre).strip(),
                    "descripcion": str(descripcion).strip() if descripcion is not None else "",
                    "responsable_id": responsable.pk,
                    "stakeholder": str(stakeholder).strip() if stakeholder is not None else "",
                    "fechaInicio": fecha_inicio,
                    "fechaLimite": fecha_limite,
                }

                if mes_planeacion:
                    data["mes_planeacion"] = mes_planeacion
                if semana_planeacion:
                    data["semana_planeacion"] = semana_planeacion

                pk_value = parse_pk(row[id_index]) if id_index is not None else None
                instance = Activity.objects.filter(pk=pk_value).first() if pk_value else None
                serializer = ActivitySerializer(
                    instance,
                    data=data,
                    partial=bool(instance),
                    context={"request": request},
                )
                if not serializer.is_valid():
                    skipped += 1
                    errors.append({"row": row_index, "error": serializer.errors})
                    continue

                save_kwargs = {}
                if instance is None and request.user and request.user.is_authenticated:
                    save_kwargs["created_by"] = request.user
                elif instance is not None and instance.created_by_id is None and request.user and request.user.is_authenticated:
                    save_kwargs["created_by"] = request.user
                serializer.save(**save_kwargs)
                if instance:
                    updated += 1
                else:
                    created += 1

            if dry_run:
                transaction.set_rollback(True)

        return Response(
            {
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "errors": errors,
                "dry_run": dry_run,
            }
        )
